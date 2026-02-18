"""
Многоформатный парсер стандартов с поддержкой специфики разных стран.
Поддерживает: PDF, Excel, CSV, JSON, XML, HTML.
Специализированные парсеры для Китая, Японии, США, Европы.
"""

import json
import logging
import re
from abc import ABC, abstractmethod
from pathlib import Path
from typing import Dict, Any, List, Optional, Tuple

logger = logging.getLogger(__name__)


class StandardsParser(ABC):
    """
    Базовый класс парсера стандартов.
    Поддерживает различные форматы: PDF, Excel, CSV, JSON, XML, HTML.
    """

    SUPPORTED_FORMATS = {
        ".pdf": "PDF",
        ".xlsx": "Excel",
        ".xls": "Excel",
        ".csv": "CSV",
        ".json": "JSON",
        ".xml": "XML",
        ".html": "HTML",
        ".htm": "HTML",
    }

    def __init__(self):
        self.detected_format = None

    @staticmethod
    def detect_format(file_path: str) -> Optional[str]:
        """
        Определить формат файла по расширению.
        
        Args:
            file_path: Путь к файлу
            
        Returns:
            Формат файла (PDF, Excel, CSV, JSON, XML, HTML) или None
        """
        path = Path(file_path)
        ext = path.suffix.lower()
        return StandardsParser.SUPPORTED_FORMATS.get(ext)

    def parse(self, file_path: str) -> Dict[str, Any]:
        """
        Парсинг файла стандарта.
        
        Args:
            file_path: Путь к файлу
            
        Returns:
            Словарь с распарсенными данными
        """
        self.detected_format = self.detect_format(file_path)
        if not self.detected_format:
            raise ValueError(f"Unsupported file format: {file_path}")

        if self.detected_format == "PDF":
            return self._parse_pdf(file_path)
        elif self.detected_format == "Excel":
            return self._parse_excel(file_path)
        elif self.detected_format == "CSV":
            return self._parse_csv(file_path)
        elif self.detected_format == "JSON":
            return self._parse_json(file_path)
        elif self.detected_format == "XML":
            return self._parse_xml(file_path)
        elif self.detected_format == "HTML":
            return self._parse_html(file_path)
        else:
            raise ValueError(f"Parser not implemented for format: {self.detected_format}")

    def _parse_pdf(self, file_path: str) -> Dict[str, Any]:
        """Парсинг PDF файла."""
        try:
            from standards.ingestion.pdf_parser import GOSTPdfParser
            parser = GOSTPdfParser()
            text = parser.extract_text(file_path)
            tables = parser.find_tables(text)
            category = parser.detect_category(text)
            return {
                "format": "PDF",
                "text": text,
                "tables": tables,
                "category": category,
                "equivalence_info": self.extract_equivalence_info(text),
            }
        except Exception as e:
            logger.error(f"Error parsing PDF {file_path}: {e}")
            return {"format": "PDF", "error": str(e)}

    def _parse_excel(self, file_path: str) -> Dict[str, Any]:
        """Парсинг Excel файла."""
        try:
            import openpyxl
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            sheets_data = {}
            text_parts = []
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                rows = []
                for row in sheet.iter_rows(values_only=True):
                    row_data = [str(cell) if cell is not None else "" for cell in row]
                    rows.append(row_data)
                    text_parts.append(" ".join(row_data))
                sheets_data[sheet_name] = rows
            
            text = "\n".join(text_parts)
            return {
                "format": "Excel",
                "sheets": sheets_data,
                "text": text,
                "equivalence_info": self.extract_equivalence_info(text),
            }
        except ImportError:
            logger.warning("openpyxl not available, cannot parse Excel")
            return {"format": "Excel", "error": "openpyxl not installed"}
        except Exception as e:
            logger.error(f"Error parsing Excel {file_path}: {e}")
            return {"format": "Excel", "error": str(e)}

    def _parse_csv(self, file_path: str) -> Dict[str, Any]:
        """Парсинг CSV файла."""
        try:
            import csv
            rows = []
            text_parts = []
            
            with open(file_path, "r", encoding="utf-8-sig") as f:
                reader = csv.reader(f)
                for row in reader:
                    rows.append(row)
                    text_parts.append(" ".join(row))
            
            text = "\n".join(text_parts)
            return {
                "format": "CSV",
                "rows": rows,
                "text": text,
                "equivalence_info": self.extract_equivalence_info(text),
            }
        except Exception as e:
            logger.error(f"Error parsing CSV {file_path}: {e}")
            return {"format": "CSV", "error": str(e)}

    def _parse_json(self, file_path: str) -> Dict[str, Any]:
        """Парсинг JSON файла."""
        try:
            with open(file_path, "r", encoding="utf-8") as f:
                data = json.load(f)
            
            # Преобразуем JSON в текст для поиска эквивалентов
            text = json.dumps(data, ensure_ascii=False)
            return {
                "format": "JSON",
                "data": data,
                "text": text,
                "equivalence_info": self.extract_equivalence_info(text),
            }
        except Exception as e:
            logger.error(f"Error parsing JSON {file_path}: {e}")
            return {"format": "JSON", "error": str(e)}

    def _parse_xml(self, file_path: str) -> Dict[str, Any]:
        """Парсинг XML файла."""
        try:
            import xml.etree.ElementTree as ET
            tree = ET.parse(file_path)
            root = tree.getroot()
            
            # Извлекаем текст из всех элементов
            text_parts = []
            for elem in root.iter():
                if elem.text:
                    text_parts.append(elem.text.strip())
            
            text = "\n".join(text_parts)
            return {
                "format": "XML",
                "root": root.tag,
                "text": text,
                "equivalence_info": self.extract_equivalence_info(text),
            }
        except Exception as e:
            logger.error(f"Error parsing XML {file_path}: {e}")
            return {"format": "XML", "error": str(e)}

    def _parse_html(self, file_path: str) -> Dict[str, Any]:
        """Парсинг HTML файла."""
        try:
            with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
                html = f.read()
            
            # Простое извлечение текста (убираем теги)
            text = re.sub(r"<[^>]+>", " ", html)
            text = re.sub(r"\s+", " ", text).strip()
            
            return {
                "format": "HTML",
                "html": html,
                "text": text,
                "equivalence_info": self.extract_equivalence_info(text),
            }
        except Exception as e:
            logger.error(f"Error parsing HTML {file_path}: {e}")
            return {"format": "HTML", "error": str(e)}

    def extract_equivalence_info(self, text: str) -> List[Dict[str, Any]]:
        """
        Извлечь информацию об аналогах из текста стандарта.
        Ищет фразы типа "equivalent to ISO 1234", "mod ISO 1234", "based on DIN 123".
        
        Args:
            text: Текст для анализа
            
        Returns:
            Список словарей с информацией об аналогах
        """
        if not text:
            return []
        
        equivalences = []
        
        # Паттерны для поиска аналогов
        patterns = [
            # "equivalent to ISO 1234"
            (r"equivalent\s+to\s+([A-Z]{2,6})\s+(\d+(?:[-.]\d+)?)", "equivalent"),
            # "mod ISO 1234" или "modified ISO 1234"
            (r"mod(?:ified)?\s+([A-Z]{2,6})\s+(\d+(?:[-.]\d+)?)", "modified"),
            # "based on DIN 123"
            (r"based\s+on\s+([A-Z]{2,6})\s+(\d+(?:[-.]\d+)?)", "based_on"),
            # "identical to GOST 123"
            (r"identical\s+to\s+([A-Z]{2,6})\s+(\d+(?:[-.]\d+)?)", "identical"),
            # "adopted from JIS 123"
            (r"adopted\s+from\s+([A-Z]{2,6})\s+(\d+(?:[-.]\d+)?)", "adopted"),
            # "harmonized with EN 123"
            (r"harmonized\s+with\s+([A-Z]{2,6})\s+(\d+(?:[-.]\d+)?)", "harmonized"),
        ]
        
        for pattern, relation_type in patterns:
            matches = re.finditer(pattern, text, re.IGNORECASE)
            for match in matches:
                source = match.group(1).upper()
                number = match.group(2)
                equivalences.append({
                    "source": source,
                    "number": number,
                    "relation": relation_type,
                    "confidence": 0.8 if relation_type == "equivalent" else 0.6,
                })
        
        return equivalences


class ChineseStandardsParser(StandardsParser):
    """
    Специализированный парсер для китайских GB стандартов.
    Работа с иероглифами, кодировками GB2312, GB18030.
    Конвертация китайских терминов в английские.
    """

    # Словарь китайских технических терминов → английские
    CHINESE_TERMS = {
        "标准": "standard",
        "螺纹": "thread",
        "公差": "tolerance",
        "配合": "fit",
        "表面粗糙度": "surface roughness",
        "材料": "material",
        "热处理": "heat treatment",
        "涂层": "coating",
        "等效": "equivalent",
        "修改": "modified",
        "基于": "based on",
    }

    def __init__(self):
        super().__init__()
        self.encoding = None

    def _detect_encoding(self, file_path: str) -> str:
        """Определить кодировку файла (GB2312, GB18030, UTF-8)."""
        encodings = ["utf-8", "gb18030", "gb2312", "gbk"]
        for enc in encodings:
            try:
                with open(file_path, "r", encoding=enc) as f:
                    f.read(1024)
                return enc
            except (UnicodeDecodeError, UnicodeError):
                continue
        return "utf-8"  # Fallback

    def _parse_csv(self, file_path: str) -> Dict[str, Any]:
        """Парсинг CSV с поддержкой китайских кодировок."""
        self.encoding = self._detect_encoding(file_path)
        try:
            import csv
            rows = []
            text_parts = []
            
            with open(file_path, "r", encoding=self.encoding) as f:
                reader = csv.reader(f)
                for row in reader:
                    rows.append(row)
                    text_parts.append(" ".join(row))
            
            text = "\n".join(text_parts)
            translated_text = self._translate_chinese_terms(text)
            
            return {
                "format": "CSV",
                "encoding": self.encoding,
                "rows": rows,
                "text": text,
                "translated_text": translated_text,
                "equivalence_info": self.extract_equivalence_info(translated_text),
            }
        except Exception as e:
            logger.error(f"Error parsing Chinese CSV {file_path}: {e}")
            return {"format": "CSV", "error": str(e)}

    def _parse_html(self, file_path: str) -> Dict[str, Any]:
        """Парсинг HTML с поддержкой китайских кодировок."""
        self.encoding = self._detect_encoding(file_path)
        try:
            with open(file_path, "r", encoding=self.encoding, errors="ignore") as f:
                html = f.read()
            
            text = re.sub(r"<[^>]+>", " ", html)
            text = re.sub(r"\s+", " ", text).strip()
            translated_text = self._translate_chinese_terms(text)
            
            return {
                "format": "HTML",
                "encoding": self.encoding,
                "html": html,
                "text": text,
                "translated_text": translated_text,
                "equivalence_info": self.extract_equivalence_info(translated_text),
            }
        except Exception as e:
            logger.error(f"Error parsing Chinese HTML {file_path}: {e}")
            return {"format": "HTML", "error": str(e)}

    def _translate_chinese_terms(self, text: str) -> str:
        """Перевести китайские технические термины в английские."""
        translated = text
        for chinese, english in self.CHINESE_TERMS.items():
            translated = translated.replace(chinese, english)
        return translated

    def extract_equivalence_info(self, text: str) -> List[Dict[str, Any]]:
        """Извлечь информацию об аналогах с учётом китайских терминов."""
        # Сначала переводим китайские термины
        translated = self._translate_chinese_terms(text)
        # Затем используем базовый метод
        equivalences = super().extract_equivalence_info(translated)
        
        # Дополнительно ищем китайские паттерны
        chinese_patterns = [
            (r"等效于\s*([A-Z]{2,6})\s*(\d+(?:[-.]\d+)?)", "equivalent"),
            (r"修改\s*([A-Z]{2,6})\s*(\d+(?:[-.]\d+)?)", "modified"),
            (r"基于\s*([A-Z]{2,6})\s*(\d+(?:[-.]\d+)?)", "based_on"),
        ]
        
        for pattern, relation_type in chinese_patterns:
            matches = re.finditer(pattern, text)
            for match in matches:
                source = match.group(1).upper()
                number = match.group(2)
                equivalences.append({
                    "source": source,
                    "number": number,
                    "relation": relation_type,
                    "confidence": 0.7,
                })
        
        return equivalences


class JapaneseStandardsParser(StandardsParser):
    """
    Специализированный парсер для японских JIS стандартов.
    Кодировки Shift-JIS, EUC-JP.
    Японские технические термины.
    """

    # Словарь японских технических терминов → английские
    JAPANESE_TERMS = {
        "規格": "standard",
        "ねじ": "thread",
        "公差": "tolerance",
        "はめあい": "fit",
        "表面粗さ": "surface roughness",
        "材料": "material",
        "熱処理": "heat treatment",
        "コーティング": "coating",
        "同等": "equivalent",
        "修正": "modified",
        "基づく": "based on",
    }

    def __init__(self):
        super().__init__()
        self.encoding = None

    def _detect_encoding(self, file_path: str) -> str:
        """Определить кодировку файла (Shift-JIS, EUC-JP, UTF-8)."""
        encodings = ["utf-8", "shift-jis", "euc-jp", "cp932"]
        for enc in encodings:
            try:
                with open(file_path, "r", encoding=enc) as f:
                    f.read(1024)
                return enc
            except (UnicodeDecodeError, UnicodeError):
                continue
        return "utf-8"

    def _parse_csv(self, file_path: str) -> Dict[str, Any]:
        """Парсинг CSV с поддержкой японских кодировок."""
        self.encoding = self._detect_encoding(file_path)
        try:
            import csv
            rows = []
            text_parts = []
            
            with open(file_path, "r", encoding=self.encoding) as f:
                reader = csv.reader(f)
                for row in reader:
                    rows.append(row)
                    text_parts.append(" ".join(row))
            
            text = "\n".join(text_parts)
            translated_text = self._translate_japanese_terms(text)
            
            return {
                "format": "CSV",
                "encoding": self.encoding,
                "rows": rows,
                "text": text,
                "translated_text": translated_text,
                "equivalence_info": self.extract_equivalence_info(translated_text),
            }
        except Exception as e:
            logger.error(f"Error parsing Japanese CSV {file_path}: {e}")
            return {"format": "CSV", "error": str(e)}

    def _parse_html(self, file_path: str) -> Dict[str, Any]:
        """Парсинг HTML с поддержкой японских кодировок."""
        self.encoding = self._detect_encoding(file_path)
        try:
            with open(file_path, "r", encoding=self.encoding, errors="ignore") as f:
                html = f.read()
            
            text = re.sub(r"<[^>]+>", " ", html)
            text = re.sub(r"\s+", " ", text).strip()
            translated_text = self._translate_japanese_terms(text)
            
            return {
                "format": "HTML",
                "encoding": self.encoding,
                "html": html,
                "text": text,
                "translated_text": translated_text,
                "equivalence_info": self.extract_equivalence_info(translated_text),
            }
        except Exception as e:
            logger.error(f"Error parsing Japanese HTML {file_path}: {e}")
            return {"format": "HTML", "error": str(e)}

    def _translate_japanese_terms(self, text: str) -> str:
        """Перевести японские технические термины в английские."""
        translated = text
        for japanese, english in self.JAPANESE_TERMS.items():
            translated = translated.replace(japanese, english)
        return translated

    def extract_equivalence_info(self, text: str) -> List[Dict[str, Any]]:
        """Извлечь информацию об аналогах с учётом японских терминов."""
        translated = self._translate_japanese_terms(text)
        equivalences = super().extract_equivalence_info(translated)
        
        # Дополнительно ищем японские паттерны
        japanese_patterns = [
            (r"同等\s*([A-Z]{2,6})\s*(\d+(?:[-.]\d+)?)", "equivalent"),
            (r"修正\s*([A-Z]{2,6})\s*(\d+(?:[-.]\d+)?)", "modified"),
        ]
        
        for pattern, relation_type in japanese_patterns:
            matches = re.finditer(pattern, text)
            for match in matches:
                source = match.group(1).upper()
                number = match.group(2)
                equivalences.append({
                    "source": source,
                    "number": number,
                    "relation": relation_type,
                    "confidence": 0.7,
                })
        
        return equivalences


class USStandardsParser(StandardsParser):
    """
    Специализированный парсер для американских стандартов.
    ANSI, ASME, ASTM, SAE.
    Дюймовая система, специфические обозначения.
    """

    # Конвертация дюймовых обозначений
    INCH_PATTERNS = [
        (r"(\d+)\s*/\s*(\d+)\s*inch", lambda m: f"{int(m.group(1)) / int(m.group(2)):.4f}"),
        (r"(\d+\.?\d*)\s*in", lambda m: f"{float(m.group(1)) * 25.4:.2f} mm"),
    ]

    def extract_equivalence_info(self, text: str) -> List[Dict[str, Any]]:
        """Извлечь информацию об аналогах с учётом американских стандартов."""
        equivalences = super().extract_equivalence_info(text)
        
        # Дополнительные паттерны для американских стандартов
        us_patterns = [
            (r"ANSI\s+([A-Z]\d+(?:[-.]\d+)?)\s+equivalent\s+to\s+([A-Z]{2,6})\s+(\d+)", "equivalent"),
            (r"ASME\s+([A-Z]\d+(?:[-.]\d+)?)\s+based\s+on\s+([A-Z]{2,6})\s+(\d+)", "based_on"),
        ]
        
        for pattern, relation_type in us_patterns:
            matches = re.finditer(pattern, text, re.IGNORECASE)
            for match in matches:
                equivalences.append({
                    "source": match.group(2).upper(),
                    "number": match.group(3),
                    "us_standard": match.group(1),
                    "relation": relation_type,
                    "confidence": 0.8,
                })
        
        return equivalences

    def _normalize_inch_units(self, text: str) -> str:
        """Нормализовать дюймовые единицы в метрические."""
        normalized = text
        for pattern, converter in self.INCH_PATTERNS:
            normalized = re.sub(pattern, converter, normalized, flags=re.IGNORECASE)
        return normalized


class EuropeanStandardsParser(StandardsParser):
    """
    Специализированный парсер для европейских стандартов.
    DIN, BS, NF, UNI - общеевропейские EN стандарты.
    Многие гармонизированы с ISO.
    """

    # EN стандарты часто гармонизированы с ISO
    EN_ISO_MAPPING = {
        "EN ISO": "ISO",
        "EN": "ISO",  # Многие EN основаны на ISO
    }

    def extract_equivalence_info(self, text: str) -> List[Dict[str, Any]]:
        """Извлечь информацию об аналогах с учётом европейских стандартов."""
        equivalences = super().extract_equivalence_info(text)
        
        # EN стандарты часто указывают базовый ISO
        en_patterns = [
            (r"EN\s+ISO\s+(\d+(?:[-.]\d+)?)", "harmonized"),
            (r"EN\s+(\d+(?:[-.]\d+)?)\s+\(ISO\s+(\d+(?:[-.]\d+)?)\)", "harmonized"),
            (r"DIN\s+EN\s+ISO\s+(\d+(?:[-.]\d+)?)", "harmonized"),
            (r"BS\s+EN\s+ISO\s+(\d+(?:[-.]\d+)?)", "harmonized"),
        ]
        
        for pattern, relation_type in en_patterns:
            matches = re.finditer(pattern, text, re.IGNORECASE)
            for match in matches:
                iso_number = match.group(1) if len(match.groups()) == 1 else match.group(2)
                equivalences.append({
                    "source": "ISO",
                    "number": iso_number,
                    "relation": relation_type,
                    "confidence": 0.9,  # EN стандарты обычно точно соответствуют ISO
                })
        
        return equivalences


# Фабрика парсеров
def get_parser_for_source(source: str) -> StandardsParser:
    """
    Получить специализированный парсер для источника стандарта.
    
    Args:
        source: Источник (GOST, GB, JIS, ANSI, DIN, BS, etc.)
        
    Returns:
        Специализированный парсер или базовый StandardsParser
    """
    source_upper = source.upper()
    
    if source_upper in ["GB", "CHINESE"]:
        return ChineseStandardsParser()
    elif source_upper in ["JIS", "JAPANESE"]:
        return JapaneseStandardsParser()
    elif source_upper in ["ANSI", "ASME", "ASTM", "SAE", "US"]:
        return USStandardsParser()
    elif source_upper in ["DIN", "BS", "NF", "UNI", "EN", "EUROPEAN"]:
        return EuropeanStandardsParser()
    else:
        return StandardsParser()
